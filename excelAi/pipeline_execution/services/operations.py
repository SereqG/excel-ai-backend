from __future__ import annotations

from copy import copy
from datetime import date, datetime
from typing import Any, Optional

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from .column_id import resolve_column_id
from .errors import PipelineValidationError


def _header_values(ws: Worksheet, *, header_row_idx: int) -> list[Any]:
    # openpyxl returns tuples of cells via iter_rows; we want current values.
    return [cell.value for cell in ws[header_row_idx]]


def apply_rename_column(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_id: str,
    new_name: Any,
) -> None:
    """
    Rename a column header.

    - new_name must be a non-empty string
    - fail on name collision (new_name already exists in header row, excluding target column)
    - column_id must match selected_sheet and header (strict)
    """
    if not isinstance(new_name, str) or not new_name.strip():
        raise PipelineValidationError("newName must be a non-empty string")
    new_name = new_name.strip()

    col_idx = resolve_column_id(
        ws,
        selected_sheet=selected_sheet,
        header_row_idx=header_row_idx,
        column_id=column_id,
    )

    header_vals = _header_values(ws, header_row_idx=header_row_idx)
    for i, val in enumerate(header_vals, start=1):
        if i == col_idx:
            continue
        if val == new_name:
            raise PipelineValidationError(f"Rename collision: header already contains '{new_name}'")

    ws.cell(row=header_row_idx, column=col_idx).value = new_name


def apply_drop_column(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_ids: Any,
) -> None:
    """
    Drop one or more columns.

    - column_ids must be a non-empty list of unique strings
    - resolve all ids on current worksheet state before deletion
    - delete from highest index to lowest to avoid shifting
    """
    if not isinstance(column_ids, list):
        raise PipelineValidationError("columnIds must be a list")
    if not column_ids:
        raise PipelineValidationError("columnIds must be a non-empty list")
    if any(not isinstance(cid, str) or not cid.strip() for cid in column_ids):
        raise PipelineValidationError("columnIds must contain non-empty strings only")
    if len(set(column_ids)) != len(column_ids):
        raise PipelineValidationError("columnIds must not contain duplicates")

    resolved: list[int] = []
    for cid in column_ids:
        resolved.append(
            resolve_column_id(
                ws,
                selected_sheet=selected_sheet,
                header_row_idx=header_row_idx,
                column_id=cid,
            )
        )

    for col_idx in sorted(set(resolved), reverse=True):
        ws.delete_cols(col_idx, 1)


def _require_non_empty_str(value: Any, *, field_name: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise PipelineValidationError(f"{field_name} must be a non-empty string")
    return value.strip()


def _parse_constant_source(source: Any) -> Any:
    if not isinstance(source, dict):
        raise PipelineValidationError("source must be an object")
    if set(source.keys()) != {"kind", "value"}:
        missing = {"kind", "value"} - set(source.keys())
        extra = set(source.keys()) - {"kind", "value"}
        raise PipelineValidationError(
            "source must have exactly keys: kind, value. "
            f"Missing={sorted(missing)} Extra={sorted(extra)}"
        )
    kind = source.get("kind")
    if kind != "constant":
        raise PipelineValidationError("source.kind must be 'constant'")
    return source.get("value")


def apply_add_column(
    ws: Worksheet,
    *,
    header_row_idx: int,
    column_name: Any,
    source: Any,
) -> None:
    """
    Append a new column as the last column.

    - column_name must be a non-empty string and must not collide with an existing header value
    - source must be an object with kind='constant' and a value to fill all data rows
    """
    column_name = _require_non_empty_str(column_name, field_name="columnName")
    fill_value = _parse_constant_source(source)

    header_vals = _header_values(ws, header_row_idx=header_row_idx)
    if column_name in header_vals:
        raise PipelineValidationError(f"Add column collision: header already contains '{column_name}'")

    template_col_idx = ws.max_column
    new_col_idx = template_col_idx + 1

    # Copy column dimension metadata from the current last column (n) to new last (n+1).
    _copy_column_dimensions(ws, from_col=template_col_idx, to_col=new_col_idx)

    # Copy styles from template cells, but overwrite values.
    for row_idx in range(1, ws.max_row + 1):
        src = ws.cell(row=row_idx, column=template_col_idx)
        dst = ws.cell(row=row_idx, column=new_col_idx)
        _copy_cell_style(src, dst)

    ws.cell(row=header_row_idx, column=new_col_idx).value = column_name
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        ws.cell(row=row_idx, column=new_col_idx).value = fill_value


def _swap_cell_contents(a, b) -> None:
    # Values
    a_val, b_val = a.value, b.value

    # Core style attributes (aligned with workbook copy logic)
    a_font, b_font = copy(a.font), copy(b.font)
    a_border, b_border = copy(a.border), copy(b.border)
    a_fill, b_fill = copy(a.fill), copy(b.fill)
    a_alignment, b_alignment = copy(a.alignment), copy(b.alignment)
    a_protection, b_protection = copy(a.protection), copy(b.protection)
    a_number_format, b_number_format = a.number_format, b.number_format

    # Other commonly used cell metadata
    a_hyperlink, b_hyperlink = a.hyperlink, b.hyperlink
    a_comment, b_comment = a.comment, b.comment

    a.value, b.value = b_val, a_val

    a.font, b.font = b_font, a_font
    a.border, b.border = b_border, a_border
    a.fill, b.fill = b_fill, a_fill
    a.alignment, b.alignment = b_alignment, a_alignment
    a.protection, b.protection = b_protection, a_protection
    a.number_format, b.number_format = b_number_format, a_number_format

    a.hyperlink, b.hyperlink = b_hyperlink, a_hyperlink
    a.comment, b.comment = b_comment, a_comment


def _copy_cell_style(src, dst) -> None:
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


_COLUMN_DIM_ATTRS: tuple[str, ...] = (
    "width",
    "hidden",
    "bestFit",
    "outlineLevel",
    "collapsed",
    "style",
)

def _set_dim_attr_if_possible(dim, attr: str, value: Any) -> None:
    """
    Some ColumnDimension attributes are read-only properties depending on openpyxl version.
    Skip any attribute that cannot be set.
    """
    try:
        setattr(dim, attr, value)
    except (AttributeError, TypeError):
        return


def _copy_column_dimensions(ws: Worksheet, *, from_col: int, to_col: int) -> None:
    letter_from = get_column_letter(from_col)
    letter_to = get_column_letter(to_col)
    dim_from = ws.column_dimensions[letter_from]
    dim_to = ws.column_dimensions[letter_to]

    for attr in _COLUMN_DIM_ATTRS:
        if hasattr(dim_from, attr) and hasattr(dim_to, attr):
            _set_dim_attr_if_possible(dim_to, attr, getattr(dim_from, attr))



def apply_reorder_columns(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_ids: Any,
) -> None:
    """
    Reorder multiple columns (including header + all data rows).

    - column_ids must be a list of >= 2 unique non-empty strings
    - both columnIds must resolve against the current worksheet state
    - columns listed are reordered to match the given order while keeping all other columns in place
    """
    if not isinstance(column_ids, list):
        raise PipelineValidationError("columnIds must be a list")
    if len(column_ids) < 2:
        raise PipelineValidationError("columnIds must contain at least two items")
    if any(not isinstance(cid, str) or not cid.strip() for cid in column_ids):
        raise PipelineValidationError("columnIds must contain non-empty strings only")
    if len(set(column_ids)) != len(column_ids):
        raise PipelineValidationError("columnIds must not contain duplicates")

    resolved: list[int] = [
        resolve_column_id(
            ws,
            selected_sheet=selected_sheet,
            header_row_idx=header_row_idx,
            column_id=cid,
        )
        for cid in column_ids
    ]
    if len(set(resolved)) != len(resolved):
        raise PipelineValidationError("columnIds must refer to distinct columns")

    # We reorder within the set of positions currently occupied by these columns,
    # leaving other columns untouched.
    dest_cols = sorted(resolved)
    max_row = ws.max_row

    def snapshot_column(col_idx: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
        cells: list[dict[str, Any]] = []
        for row_idx in range(1, max_row + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            cells.append(
                {
                    "value": c.value,
                    "font": copy(c.font),
                    "border": copy(c.border),
                    "fill": copy(c.fill),
                    "alignment": copy(c.alignment),
                    "protection": copy(c.protection),
                    "number_format": c.number_format,
                    "hyperlink": copy(c.hyperlink) if c.hyperlink else None,
                    "comment": copy(c.comment) if c.comment else None,
                }
            )

        letter = get_column_letter(col_idx)
        dim = ws.column_dimensions[letter]
        dim_snap: dict[str, Any] = {}
        for attr in _COLUMN_DIM_ATTRS:
            if hasattr(dim, attr):
                dim_snap[attr] = getattr(dim, attr)
        return cells, dim_snap

    def apply_column_snapshot(dest_col: int, *, cells: list[dict[str, Any]], dim_snap: dict[str, Any]) -> None:
        for row_idx, snap in enumerate(cells, start=1):
            c = ws.cell(row=row_idx, column=dest_col)
            c.value = snap["value"]
            c.font = snap["font"]
            c.border = snap["border"]
            c.fill = snap["fill"]
            c.alignment = snap["alignment"]
            c.protection = snap["protection"]
            c.number_format = snap["number_format"]
            c.hyperlink = snap["hyperlink"]
            c.comment = snap["comment"]

        letter = get_column_letter(dest_col)
        dim = ws.column_dimensions[letter]
        for attr, val in dim_snap.items():
            if hasattr(dim, attr):
                _set_dim_attr_if_possible(dim, attr, val)

    snapshots = [snapshot_column(col) for col in resolved]
    for dest_col, (cells, dim_snap) in zip(dest_cols, snapshots):
        apply_column_snapshot(dest_col, cells=cells, dim_snap=dim_snap)


def _filter_non_empty_values(values: list[Any]) -> list[Any]:
    return [v for v in values if v is not None and v != ""]


def _is_numeric_value(value: Any) -> bool:
    if isinstance(value, (int, float)):
        return True

    if isinstance(value, str):
        try:
            float(value.replace(",", ""))
            return True
        except (ValueError, AttributeError):
            return False

    return False


def _is_date_value(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True

    if isinstance(value, str):
        for date_format in _DATE_INPUT_STRPTIME_FORMATS:
            try:
                datetime.strptime(value, date_format)
                return True
            except (ValueError, TypeError):
                continue

    return False


_ALLOWED_DATE_OUTPUT_FORMATS: set[str] = {
    "YYYY/MM/DD",
    "DD/MM/YYYY",
    "YYYY.MM.DD",
    "DD.MM.YYYY",
    "YYYY-MM-DD",
    "DD-MM-YYYY",
}

# Keep existing supported input parsing formats and extend with the requested ones.
_DATE_INPUT_STRPTIME_FORMATS: tuple[str, ...] = (
    "%Y/%m/%d",
    "%d/%m/%Y",
    "%Y.%m.%d",
    "%d.%m.%Y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    # legacy / already supported
    "%m/%d/%Y",
)

_DATE_OUTPUT_FORMAT_TO_EXCEL_NUMBER_FORMAT: dict[str, str] = {
    "YYYY/MM/DD": "yyyy/mm/dd",
    "DD/MM/YYYY": "dd/mm/yyyy",
    "YYYY.MM.DD": "yyyy.mm.dd",
    "DD.MM.YYYY": "dd.mm.yyyy",
    "YYYY-MM-DD": "yyyy-mm-dd",
    "DD-MM-YYYY": "dd-mm-yyyy",
}


def _require_date_dtype_and_resolve(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_id: str,
) -> int:
    col_idx = resolve_column_id(
        ws,
        selected_sheet=selected_sheet,
        header_row_idx=header_row_idx,
        column_id=column_id,
    )

    values = _collect_column_values(ws, header_row_idx=header_row_idx, col_idx=col_idx)
    non_empty = _filter_non_empty_values(values)
    if not non_empty:
        raise PipelineValidationError(
            f"ColumnId={column_id} must contain at least one non-empty value to parse as date"
        )

    inferred = _determine_column_type(values)
    if inferred != "date":
        raise PipelineValidationError(f"ColumnId={column_id} must be date. Inferred type={inferred}")

    return col_idx


def _parse_date_cell_value(value: Any) -> date:
    if isinstance(value, (datetime, date)):
        return value

    if isinstance(value, str):
        s = value.strip()
        if not s:
            raise ValueError("Empty string")
        for fmt in _DATE_INPUT_STRPTIME_FORMATS:
            try:
                return datetime.strptime(s, fmt)
            except (ValueError, TypeError):
                continue
        raise ValueError(f"Unsupported date string format: {s!r}")

    raise ValueError(f"Unsupported date cell type: {type(value).__name__}")


def apply_parse_date(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_ids: Any,
    output_format: Any,
) -> None:
    """
    Parse date-like columns and apply a consistent Excel date number format.

    Params:
    - column_ids: list of columnId strings
    - output_format: one of the supported 'YYYY/MM/DD' style strings
    """
    if not isinstance(column_ids, list):
        raise PipelineValidationError("columnIds must be a list")
    if not column_ids:
        raise PipelineValidationError("columnIds must be a non-empty list")
    if any(not isinstance(cid, str) or not cid.strip() for cid in column_ids):
        raise PipelineValidationError("columnIds must contain non-empty strings only")
    if len(set(column_ids)) != len(column_ids):
        raise PipelineValidationError("columnIds must not contain duplicates")

    output_format_str = _require_non_empty_str(output_format, field_name="outputFormat")
    if output_format_str not in _ALLOWED_DATE_OUTPUT_FORMATS:
        raise PipelineValidationError(
            f"outputFormat must be one of {sorted(_ALLOWED_DATE_OUTPUT_FORMATS)}"
        )
    excel_number_format = _DATE_OUTPUT_FORMAT_TO_EXCEL_NUMBER_FORMAT[output_format_str]

    for column_id in column_ids:
        column_id = column_id.strip()
        col_idx = _require_date_dtype_and_resolve(
            ws,
            header_row_idx=header_row_idx,
            selected_sheet=selected_sheet,
            column_id=column_id,
        )

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None or val == "":
                continue

            try:
                parsed = _parse_date_cell_value(val)
            except Exception as exc:
                raise PipelineValidationError(
                    f"Failed to parse date in ColumnId={column_id} at row={row_idx}: "
                    f"value={val!r} (type={type(val).__name__}). {exc}"
                )

            cell.value = parsed
            cell.number_format = excel_number_format


def _count_numeric_values(values: list[Any]) -> int:
    count = 0
    for value in values:
        if _is_numeric_value(value):
            count += 1
    return count


def _count_date_values(values: list[Any]) -> int:
    count = 0
    for value in values:
        if _is_date_value(value):
            count += 1
    return count


def _count_boolean_values(values: list[Any]) -> int:
    return sum(1 for v in values if isinstance(v, bool))


def _determine_type_from_counts(total: int, bool_count: int, date_count: int, number_count: int) -> str:
    if bool_count == total:
        return "boolean"

    if date_count / total >= 0.5:
        return "date"

    if number_count / total >= 0.8:
        return "number"

    return "string"


def _determine_column_type(values: list[Any]) -> Optional[str]:
    if not values:
        return None

    non_empty_values = _filter_non_empty_values(values)
    if not non_empty_values:
        return None

    bool_count = _count_boolean_values(non_empty_values)
    date_count = _count_date_values(non_empty_values)
    number_count = _count_numeric_values(non_empty_values)
    total = len(non_empty_values)

    return _determine_type_from_counts(total, bool_count, date_count, number_count)


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(cell is None or cell == "" for cell in row)


def _collect_column_values(ws: Worksheet, *, header_row_idx: int, col_idx: int) -> list[Any]:
    values: list[Any] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if _is_empty_row(row):
            continue
        values.append(row[col_idx - 1] if (col_idx - 1) < len(row) else None)
    return values


def _require_text_dtype_and_resolve(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_id: str,
) -> int:
    col_idx = resolve_column_id(
        ws,
        selected_sheet=selected_sheet,
        header_row_idx=header_row_idx,
        column_id=column_id,
    )
    inferred = _determine_column_type(_collect_column_values(ws, header_row_idx=header_row_idx, col_idx=col_idx))
    if inferred != "string":
        raise PipelineValidationError(
            f"ColumnId={column_id} must be text (string). Inferred type={inferred}"
        )
    return col_idx


_ALLOWED_TEXT_CASES: set[str] = {"lowercase", "uppercase", "sentence_case", "title_case"}


def _parse_normalize_targets(targets_raw: Any) -> list[dict[str, str]]:
    if not isinstance(targets_raw, list):
        raise PipelineValidationError("targets must be a list")
    if not targets_raw:
        raise PipelineValidationError("targets must be a non-empty list")

    targets: list[dict[str, str]] = []
    for i, raw in enumerate(targets_raw):
        if not isinstance(raw, dict):
            raise PipelineValidationError(f"targets[{i}] must be an object")
        if set(raw.keys()) != {"columnId", "textCase"}:
            missing = {"columnId", "textCase"} - set(raw.keys())
            extra = set(raw.keys()) - {"columnId", "textCase"}
            raise PipelineValidationError(
                f"targets[{i}] must have exactly keys: columnId, textCase. "
                f"Missing={sorted(missing)} Extra={sorted(extra)}"
            )

        column_id = raw.get("columnId")
        text_case = raw.get("textCase")

        if not isinstance(column_id, str) or not column_id.strip():
            raise PipelineValidationError(f"targets[{i}].columnId must be a non-empty string")
        if not isinstance(text_case, str):
            raise PipelineValidationError(f"targets[{i}].textCase must be a string")
        if text_case not in _ALLOWED_TEXT_CASES:
            raise PipelineValidationError(
                f"targets[{i}].textCase must be one of {sorted(_ALLOWED_TEXT_CASES)}"
            )

        targets.append({"columnId": column_id.strip(), "textCase": text_case})

    return targets


def _sentence_case(s: str) -> str:
    lowered = s.lower()
    for idx, ch in enumerate(lowered):
        if ch.lower() != ch.upper():
            return lowered[:idx] + ch.upper() + lowered[idx + 1 :]
    return lowered


def _apply_text_case(s: str, *, text_case: str) -> str:
    if text_case == "lowercase":
        return s.lower()
    if text_case == "uppercase":
        return s.upper()
    if text_case == "sentence_case":
        return _sentence_case(s)
    if text_case == "title_case":
        return s.title()
    raise PipelineValidationError(f"Unsupported textCase: {text_case}")


def apply_normalize_case(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    targets: Any,
) -> None:
    parsed_targets = _parse_normalize_targets(targets)

    for target in parsed_targets:
        column_id = target["columnId"]
        text_case = target["textCase"]

        col_idx = _require_text_dtype_and_resolve(
            ws,
            header_row_idx=header_row_idx,
            selected_sheet=selected_sheet,
            column_id=column_id,
        )

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None or val == "":
                continue
            s = val if isinstance(val, str) else str(val)
            cell.value = _apply_text_case(s, text_case=text_case)


def apply_replace_text(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_id: Any,
    find_text: Any,
    replace_text: Any,
) -> None:
    if not isinstance(column_id, str) or not column_id.strip():
        raise PipelineValidationError("columnId must be a non-empty string")
    if not isinstance(find_text, str) or find_text == "":
        raise PipelineValidationError("findText must be a non-empty string")
    if not isinstance(replace_text, str):
        raise PipelineValidationError("replaceText must be a string")

    column_id = column_id.strip()

    col_idx = _require_text_dtype_and_resolve(
        ws,
        header_row_idx=header_row_idx,
        selected_sheet=selected_sheet,
        column_id=column_id,
    )

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is None or val == "":
            continue
        s = val if isinstance(val, str) else str(val)
        cell.value = s.replace(find_text, replace_text)


_FILTER_ROWS_ACTIONS: set[str] = {"keep", "drop"}
_FILTER_ROWS_WHEN_OPS: set[str] = {"and", "or"}

_FILTER_ROWS_OPERATORS_BY_TYPE: dict[str, set[str]] = {
    "text": {
        "equals",
        "not_equals",
        "contains",
        "starts_with",
        "ends_with",
        "is_empty",
        "is_not_empty",
    },
    "number": {"eq", "ne", "gt", "gte", "lt", "lte", "between", "is_empty", "is_not_empty"},
    "date": {"on", "before", "after", "between", "is_empty", "is_not_empty"},
    "boolean": {"is_true", "is_false", "is_empty", "is_not_empty"},
}


def _require_dict(value: Any, *, field_name: str) -> dict[str, Any]:
    if not isinstance(value, dict):
        raise PipelineValidationError(f"{field_name} must be an object")
    return value


def _require_list(value: Any, *, field_name: str) -> list[Any]:
    if not isinstance(value, list):
        raise PipelineValidationError(f"{field_name} must be a list")
    return value


def _require_exact_keys(obj: dict[str, Any], *, expected_keys: set[str], field_name: str) -> None:
    actual_keys = set(obj.keys())
    if actual_keys == expected_keys:
        return
    missing = expected_keys - actual_keys
    extra = actual_keys - expected_keys
    raise PipelineValidationError(
        f"{field_name} must have exactly keys {sorted(expected_keys)}. "
        f"Missing={sorted(missing)} Extra={sorted(extra)}"
    )


def _parse_number_value(value: Any, *, field_name: str) -> float:
    if isinstance(value, bool):
        # bool is a subclass of int; treat it as invalid for numeric comparisons
        raise PipelineValidationError(f"{field_name} must be a number")
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip()
        if not s:
            raise PipelineValidationError(f"{field_name} must be a number")
        try:
            return float(s.replace(",", ""))
        except ValueError:
            raise PipelineValidationError(f"{field_name} must be a number")
    raise PipelineValidationError(f"{field_name} must be a number")


def _parse_cell_number(value: Any, *, column_id: str, row_idx: int) -> float:
    try:
        return _parse_number_value(value, field_name="cellValue")
    except PipelineValidationError as exc:
        raise PipelineValidationError(
            f"Failed to parse number in ColumnId={column_id} at row={row_idx}: "
            f"value={value!r} (type={type(value).__name__}). {exc}"
        )


def _parse_date_value(value: Any, *, field_name: str) -> date:
    if value is None or value == "":
        raise PipelineValidationError(f"{field_name} must be a date")

    try:
        parsed = _parse_date_cell_value(value)
    except Exception as exc:
        raise PipelineValidationError(f"{field_name} must be a date. {exc}")

    if isinstance(parsed, datetime):
        return parsed.date()
    if isinstance(parsed, date):
        return parsed
    raise PipelineValidationError(f"{field_name} must be a date")


def _parse_cell_date(value: Any, *, column_id: str, row_idx: int) -> date:
    try:
        return _parse_date_value(value, field_name="cellValue")
    except PipelineValidationError as exc:
        raise PipelineValidationError(
            f"Failed to parse date in ColumnId={column_id} at row={row_idx}: "
            f"value={value!r} (type={type(value).__name__}). {exc}"
        )


def _infer_column_type_cached(
    ws: Worksheet,
    *,
    header_row_idx: int,
    col_idx: int,
    cache: dict[int, Optional[str]],
) -> Optional[str]:
    if col_idx in cache:
        return cache[col_idx]
    inferred = _determine_column_type(_collect_column_values(ws, header_row_idx=header_row_idx, col_idx=col_idx))
    cache[col_idx] = inferred
    return inferred


def apply_filter_rows(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    default_action: Any,
    rules: Any,
) -> None:
    """
    Filter worksheet rows (keep/drop) based on ordered rules.

    Semantics:
    - evaluate all rules in order (last matching rule wins)
    - if no rule matches, apply default_action
    - action is applied when the rule's `when` evaluates to True
    """

    def require_action(value: Any, *, field_name: str) -> str:
        if not isinstance(value, str):
            raise PipelineValidationError(f"{field_name} must be a string")
        s = value.strip()
        if s not in _FILTER_ROWS_ACTIONS:
            raise PipelineValidationError(f"{field_name} must be one of {sorted(_FILTER_ROWS_ACTIONS)}")
        return s

    default_action_str = require_action(default_action, field_name="defaultAction")

    rules_list = _require_list(rules, field_name="rules")
    if not rules_list:
        raise PipelineValidationError("rules must be a non-empty list")

    col_type_cache: dict[int, Optional[str]] = {}

    compiled_rules: list[dict[str, Any]] = []
    for rule_idx, raw_rule in enumerate(rules_list):
        rule = _require_dict(raw_rule, field_name=f"rules[{rule_idx}]")
        _require_exact_keys(rule, expected_keys={"action", "when"}, field_name=f"rules[{rule_idx}]")

        action_str = require_action(rule.get("action"), field_name=f"rules[{rule_idx}].action")
        when_raw = _require_dict(rule.get("when"), field_name=f"rules[{rule_idx}].when")
        _require_exact_keys(when_raw, expected_keys={"op", "conditions"}, field_name=f"rules[{rule_idx}].when")

        op_raw = when_raw.get("op")
        if not isinstance(op_raw, str):
            raise PipelineValidationError(f"rules[{rule_idx}].when.op must be a string")
        op_str = op_raw.strip()
        if op_str not in _FILTER_ROWS_WHEN_OPS:
            raise PipelineValidationError(
                f"rules[{rule_idx}].when.op must be one of {sorted(_FILTER_ROWS_WHEN_OPS)}"
            )

        conditions_list = _require_list(
            when_raw.get("conditions"), field_name=f"rules[{rule_idx}].when.conditions"
        )
        if not conditions_list:
            raise PipelineValidationError(f"rules[{rule_idx}].when.conditions must be a non-empty list")

        compiled_conditions: list[dict[str, Any]] = []
        for cond_idx, raw_cond in enumerate(conditions_list):
            cond = _require_dict(raw_cond, field_name=f"rules[{rule_idx}].when.conditions[{cond_idx}]")
            allowed_keys = {"type", "columnId", "operator", "value", "value2"}
            extra_keys = set(cond.keys()) - allowed_keys
            if extra_keys:
                raise PipelineValidationError(
                    f"rules[{rule_idx}].when.conditions[{cond_idx}] contains unknown keys: "
                    f"{sorted(extra_keys)}"
                )

            type_raw = cond.get("type")
            if not isinstance(type_raw, str):
                raise PipelineValidationError(
                    f"rules[{rule_idx}].when.conditions[{cond_idx}].type must be a string"
                )
            cond_type = type_raw.strip()
            if cond_type not in _FILTER_ROWS_OPERATORS_BY_TYPE:
                raise PipelineValidationError(
                    f"rules[{rule_idx}].when.conditions[{cond_idx}].type must be one of "
                    f"{sorted(_FILTER_ROWS_OPERATORS_BY_TYPE.keys())}"
                )

            column_id_raw = cond.get("columnId")
            if not isinstance(column_id_raw, str) or not column_id_raw.strip():
                raise PipelineValidationError(
                    f"rules[{rule_idx}].when.conditions[{cond_idx}].columnId must be a non-empty string"
                )
            column_id = column_id_raw.strip()
            col_idx = resolve_column_id(
                ws,
                selected_sheet=selected_sheet,
                header_row_idx=header_row_idx,
                column_id=column_id,
            )

            operator_raw = cond.get("operator")
            if not isinstance(operator_raw, str):
                raise PipelineValidationError(
                    f"rules[{rule_idx}].when.conditions[{cond_idx}].operator must be a string"
                )
            operator = operator_raw.strip()
            allowed_ops = _FILTER_ROWS_OPERATORS_BY_TYPE[cond_type]
            if operator not in allowed_ops:
                raise PipelineValidationError(
                    f"rules[{rule_idx}].when.conditions[{cond_idx}].operator must be one of "
                    f"{sorted(allowed_ops)} for type={cond_type}"
                )

            inferred = _infer_column_type_cached(
                ws, header_row_idx=header_row_idx, col_idx=col_idx, cache=col_type_cache
            )
            expected_inferred: Optional[str]
            if cond_type == "text":
                expected_inferred = "string"
            else:
                expected_inferred = cond_type

            if inferred is not None and inferred != expected_inferred:
                raise PipelineValidationError(
                    f"ColumnId={column_id} must be {cond_type}. Inferred type={inferred}"
                )

            compiled: dict[str, Any] = {
                "type": cond_type,
                "columnId": column_id,
                "colIdx": col_idx,
                "operator": operator,
            }

            # Validate and pre-parse constants
            if operator == "between":
                if "value" not in cond or "value2" not in cond:
                    raise PipelineValidationError(
                        f"rules[{rule_idx}].when.conditions[{cond_idx}] must include value and value2 for between"
                    )
                raw_min = cond.get("value")
                raw_max = cond.get("value2")
                if cond_type == "number":
                    min_v = _parse_number_value(raw_min, field_name="value")
                    max_v = _parse_number_value(raw_max, field_name="value2")
                    if min_v > max_v:
                        raise PipelineValidationError("between requires value <= value2")
                    compiled["value"] = min_v
                    compiled["value2"] = max_v
                elif cond_type == "date":
                    min_d = _parse_date_value(raw_min, field_name="value")
                    max_d = _parse_date_value(raw_max, field_name="value2")
                    if min_d > max_d:
                        raise PipelineValidationError("between requires value <= value2")
                    compiled["value"] = min_d
                    compiled["value2"] = max_d
                else:
                    raise PipelineValidationError("between operator is only supported for number/date")
            elif operator in {"is_empty", "is_not_empty", "is_true", "is_false"}:
                # no constants needed
                pass
            else:
                if "value" not in cond:
                    raise PipelineValidationError(
                        f"rules[{rule_idx}].when.conditions[{cond_idx}] must include value"
                    )
                raw_value = cond.get("value")
                if cond_type == "number":
                    compiled["value"] = _parse_number_value(raw_value, field_name="value")
                elif cond_type == "date":
                    compiled["value"] = _parse_date_value(raw_value, field_name="value")
                elif cond_type == "text":
                    if raw_value is None:
                        raise PipelineValidationError("value must be a string")
                    compiled["value"] = raw_value if isinstance(raw_value, str) else str(raw_value)
                elif cond_type == "boolean":
                    # boolean operators don't use value currently
                    pass

            compiled_conditions.append(compiled)

        compiled_rules.append({"action": action_str, "op": op_str, "conditions": compiled_conditions})

    def is_empty_cell(v: Any) -> bool:
        return v is None or v == ""

    def eval_condition(*, row_idx: int, compiled_condition: dict[str, Any]) -> bool:
        cond_type = compiled_condition["type"]
        operator = compiled_condition["operator"]
        col_idx = compiled_condition["colIdx"]
        column_id = compiled_condition["columnId"]

        cell_value = ws.cell(row=row_idx, column=col_idx).value

        if operator == "is_empty":
            return is_empty_cell(cell_value)
        if operator == "is_not_empty":
            return not is_empty_cell(cell_value)

        if is_empty_cell(cell_value):
            return False

        if cond_type == "text":
            s = cell_value if isinstance(cell_value, str) else str(cell_value)
            target = compiled_condition.get("value", "")
            if operator == "equals":
                return s == target
            if operator == "not_equals":
                return s != target
            if operator == "contains":
                return target in s
            if operator == "starts_with":
                return s.startswith(target)
            if operator == "ends_with":
                return s.endswith(target)
            raise PipelineValidationError(f"Unsupported text operator: {operator}")

        if cond_type == "number":
            num = _parse_cell_number(cell_value, column_id=column_id, row_idx=row_idx)
            if operator == "eq":
                return num == compiled_condition["value"]
            if operator == "ne":
                return num != compiled_condition["value"]
            if operator == "gt":
                return num > compiled_condition["value"]
            if operator == "gte":
                return num >= compiled_condition["value"]
            if operator == "lt":
                return num < compiled_condition["value"]
            if operator == "lte":
                return num <= compiled_condition["value"]
            if operator == "between":
                return compiled_condition["value"] <= num <= compiled_condition["value2"]
            raise PipelineValidationError(f"Unsupported number operator: {operator}")

        if cond_type == "date":
            d = _parse_cell_date(cell_value, column_id=column_id, row_idx=row_idx)
            if operator == "on":
                return d == compiled_condition["value"]
            if operator == "before":
                return d < compiled_condition["value"]
            if operator == "after":
                return d > compiled_condition["value"]
            if operator == "between":
                return compiled_condition["value"] <= d <= compiled_condition["value2"]
            raise PipelineValidationError(f"Unsupported date operator: {operator}")

        if cond_type == "boolean":
            if not isinstance(cell_value, bool):
                raise PipelineValidationError(
                    f"Failed to parse boolean in ColumnId={column_id} at row={row_idx}: "
                    f"value={cell_value!r} (type={type(cell_value).__name__})"
                )
            if operator == "is_true":
                return cell_value is True
            if operator == "is_false":
                return cell_value is False
            raise PipelineValidationError(f"Unsupported boolean operator: {operator}")

        raise PipelineValidationError(f"Unsupported condition type: {cond_type}")

    def eval_rule(*, row_idx: int, rule: dict[str, Any]) -> bool:
        op_str = rule["op"]
        conditions = rule["conditions"]
        if op_str == "and":
            for c in conditions:
                if not eval_condition(row_idx=row_idx, compiled_condition=c):
                    return False
            return True
        if op_str == "or":
            for c in conditions:
                if eval_condition(row_idx=row_idx, compiled_condition=c):
                    return True
            return False
        raise PipelineValidationError(f"Unsupported when.op: {op_str}")

    rows_to_drop: list[int] = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        row_action: Optional[str] = None
        for rule in compiled_rules:
            if eval_rule(row_idx=row_idx, rule=rule):
                row_action = rule["action"]
        final_action = row_action or default_action_str
        if final_action == "drop":
            rows_to_drop.append(row_idx)

    for row_idx in sorted(rows_to_drop, reverse=True):
        ws.delete_rows(row_idx, 1)

