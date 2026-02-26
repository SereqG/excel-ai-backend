from __future__ import annotations

from typing import Any, Optional

from openpyxl.worksheet.worksheet import Worksheet

from ..column_id import resolve_column_id
from ..errors import PipelineValidationError
from .common import _collect_column_values, _determine_column_type


def _require_text_dtype_and_resolve(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_id: str,
) -> int:
    col_idx = resolve_column_id(
        ws,
        selected_sheet=selected_sheet,
        header_row_idx=header_row_idx,
        column_id=column_id,
    )
    inferred = _determine_column_type(_collect_column_values(ws, header_row_idx=header_row_idx, col_idx=col_idx))
    if inferred != "string":
        raise PipelineValidationError(
            f"ColumnId={column_id} must be text (string). Inferred type={inferred}"
        )
    return col_idx


_ALLOWED_TEXT_CASES: set[str] = {"lowercase", "uppercase", "sentence_case", "title_case"}


def _parse_normalize_targets(targets_raw: Any) -> list[dict[str, str]]:
    if not isinstance(targets_raw, list):
        raise PipelineValidationError("targets must be a list")
    if not targets_raw:
        raise PipelineValidationError("targets must be a non-empty list")

    targets: list[dict[str, str]] = []
    for i, raw in enumerate(targets_raw):
        if not isinstance(raw, dict):
            raise PipelineValidationError(f"targets[{i}] must be an object")
        if set(raw.keys()) != {"columnId", "textCase"}:
            missing = {"columnId", "textCase"} - set(raw.keys())
            extra = set(raw.keys()) - {"columnId", "textCase"}
            raise PipelineValidationError(
                f"targets[{i}] must have exactly keys: columnId, textCase. "
                f"Missing={sorted(missing)} Extra={sorted(extra)}"
            )

        column_id = raw.get("columnId")
        text_case = raw.get("textCase")

        if not isinstance(column_id, str) or not column_id.strip():
            raise PipelineValidationError(f"targets[{i}].columnId must be a non-empty string")
        if not isinstance(text_case, str):
            raise PipelineValidationError(f"targets[{i}].textCase must be a string")
        if text_case not in _ALLOWED_TEXT_CASES:
            raise PipelineValidationError(
                f"targets[{i}].textCase must be one of {sorted(_ALLOWED_TEXT_CASES)}"
            )

        targets.append({"columnId": column_id.strip(), "textCase": text_case})

    return targets


def _sentence_case(s: str) -> str:
    lowered = s.lower()
    for idx, ch in enumerate(lowered):
        if ch.lower() != ch.upper():
            return lowered[:idx] + ch.upper() + lowered[idx + 1 :]
    return lowered


def _apply_text_case(s: str, *, text_case: str) -> str:
    if text_case == "lowercase":
        return s.lower()
    if text_case == "uppercase":
        return s.upper()
    if text_case == "sentence_case":
        return _sentence_case(s)
    if text_case == "title_case":
        return s.title()
    raise PipelineValidationError(f"Unsupported textCase: {text_case}")


def apply_normalize_case(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    targets: Any,
) -> None:
    parsed_targets = _parse_normalize_targets(targets)

    for target in parsed_targets:
        column_id = target["columnId"]
        text_case = target["textCase"]

        col_idx = _require_text_dtype_and_resolve(
            ws,
            header_row_idx=header_row_idx,
            selected_sheet=selected_sheet,
            column_id=column_id,
        )

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is None or val == "":
                continue
            s = val if isinstance(val, str) else str(val)
            cell.value = _apply_text_case(s, text_case=text_case)


def apply_replace_text(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    column_id: Any,
    find_text: Any,
    replace_text: Any,
) -> None:
    if not isinstance(column_id, str) or not column_id.strip():
        raise PipelineValidationError("columnId must be a non-empty string")
    if not isinstance(find_text, str) or find_text == "":
        raise PipelineValidationError("findText must be a non-empty string")
    if not isinstance(replace_text, str):
        raise PipelineValidationError("replaceText must be a string")

    column_id = column_id.strip()

    col_idx = _require_text_dtype_and_resolve(
        ws,
        header_row_idx=header_row_idx,
        selected_sheet=selected_sheet,
        column_id=column_id,
    )

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        val = cell.value
        if val is None or val == "":
            continue
        s = val if isinstance(val, str) else str(val)
        cell.value = s.replace(find_text, replace_text)

