from __future__ import annotations

from copy import copy
from datetime import date, datetime
from typing import Any, Optional

from openpyxl.worksheet.worksheet import Worksheet

from ..column_id import resolve_column_id
from ..errors import PipelineValidationError
from .common import (
    _infer_column_type_cached,
    _is_empty_row,
    _parse_date_cell_value,
    _parse_number_value,
    _set_dim_attr_if_possible,
)


_ROW_DIM_ATTRS: tuple[str, ...] = (
    "height",
    "hidden",
    "outlineLevel",
    "collapsed",
    "style",
)


def _parse_sort_rows_criteria(sorts_raw: Any) -> list[dict[str, str]]:
    if not isinstance(sorts_raw, list):
        raise PipelineValidationError("sorts must be a list")
    if not sorts_raw:
        raise PipelineValidationError("sorts must be a non-empty list")

    parsed: list[dict[str, str]] = []
    for i, raw in enumerate(sorts_raw):
        if not isinstance(raw, dict):
            raise PipelineValidationError(f"sorts[{i}] must be an object")

        if set(raw.keys()) != {"columnId", "direction"}:
            missing = {"columnId", "direction"} - set(raw.keys())
            extra = set(raw.keys()) - {"columnId", "direction"}
            raise PipelineValidationError(
                "Each sort criteria must have exactly keys: columnId, direction. "
                f"Missing={sorted(missing)} Extra={sorted(extra)}"
            )

        column_id_raw = raw.get("columnId")
        direction_raw = raw.get("direction")

        if not isinstance(column_id_raw, str):
            raise PipelineValidationError(f"sorts[{i}].columnId must be a non-empty string")
        if not isinstance(direction_raw, str):
            raise PipelineValidationError(f"sorts[{i}].direction must be a string")

        column_id = column_id_raw.strip()
        if not column_id:
            raise PipelineValidationError(f"sorts[{i}].columnId must be a non-empty string")

        direction = direction_raw.strip()
        if direction not in {"asc", "desc"}:
            raise PipelineValidationError(f"sorts[{i}].direction must be 'asc' or 'desc'")

        parsed.append({"columnId": column_id, "direction": direction})

    return parsed


def _snapshot_row(ws: Worksheet, *, row_idx: int, max_col: int) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for col_idx in range(1, max_col + 1):
        c = ws.cell(row=row_idx, column=col_idx)
        cells.append(
            {
                "value": c.value,
                "font": copy(c.font),
                "border": copy(c.border),
                "fill": copy(c.fill),
                "alignment": copy(c.alignment),
                "protection": copy(c.protection),
                "number_format": c.number_format,
                "hyperlink": copy(c.hyperlink) if c.hyperlink else None,
                "comment": copy(c.comment) if c.comment else None,
            }
        )

    dim = ws.row_dimensions[row_idx]
    dim_snap: dict[str, Any] = {}
    for attr in _ROW_DIM_ATTRS:
        if hasattr(dim, attr):
            dim_snap[attr] = getattr(dim, attr)

    return cells, dim_snap


def _apply_row_snapshot(
    ws: Worksheet,
    *,
    dest_row_idx: int,
    cells: list[dict[str, Any]],
    dim_snap: dict[str, Any],
) -> None:
    for col_idx, snap in enumerate(cells, start=1):
        c = ws.cell(row=dest_row_idx, column=col_idx)
        c.value = snap["value"]
        c.font = snap["font"]
        c.border = snap["border"]
        c.fill = snap["fill"]
        c.alignment = snap["alignment"]
        c.protection = snap["protection"]
        c.number_format = snap["number_format"]
        c.hyperlink = snap["hyperlink"]
        c.comment = snap["comment"]

    dim = ws.row_dimensions[dest_row_idx]
    for attr, val in dim_snap.items():
        if hasattr(dim, attr):
            _set_dim_attr_if_possible(dim, attr, val)


def _make_sort_key(
    value: Any,
    *,
    inferred_type: Optional[str],
    reverse: bool,
) -> tuple[int, int, Any]:
    """
    Produce a sort key that:
    - keeps blanks direction-aware (asc: blanks last, desc: blanks first)
    - avoids TypeError by separating parseable vs fallback values
    """
    is_blank = value is None or value == ""
    blank_flag = 1 if is_blank else 0

    # Keep all blanks equal so stable sort preserves original order among blanks.
    if is_blank:
        return (blank_flag, 0, 0)

    inferred = inferred_type or "string"

    type_rank = 0
    sort_value: Any

    if inferred == "number":
        try:
            sort_value = _parse_number_value(value, field_name="cellValue")
        except Exception:
            type_rank = 1
            sort_value = value if isinstance(value, str) else str(value)
    elif inferred == "date":
        try:
            parsed = _parse_date_cell_value(value)
            if isinstance(parsed, datetime):
                sort_value = parsed.date()
            elif isinstance(parsed, date):
                sort_value = parsed
            else:
                raise ValueError("Unsupported date parse result")
        except Exception:
            type_rank = 1
            sort_value = value if isinstance(value, str) else str(value)
    elif inferred == "boolean":
        if isinstance(value, bool):
            sort_value = 1 if value else 0
        else:
            type_rank = 1
            sort_value = value if isinstance(value, str) else str(value)
    else:
        sort_value = value if isinstance(value, str) else str(value)

    type_rank_component = -type_rank if reverse else type_rank
    return (blank_flag, type_rank_component, sort_value)


def _get_sort_rows_dimensions(ws: Worksheet, *, header_row_idx: int) -> Optional[tuple[int, int]]:
    max_row = ws.max_row
    if max_row <= header_row_idx + 1:
        return None

    max_col = ws.max_column
    if max_col <= 0:
        return None

    return max_row, max_col


def apply_sort_rows(
    ws: Worksheet,
    *,
    header_row_idx: int,
    selected_sheet: str,
    sorts: Any,
) -> None:
    """
    Sort worksheet rows based on ordered criteria.

    Params:
    - sorts: ordered list of { columnId, direction } where direction is 'asc' or 'desc'
    """
    parsed_sorts = _parse_sort_rows_criteria(sorts)

    dims = _get_sort_rows_dimensions(ws, header_row_idx=header_row_idx)
    if dims is None:
        return
    max_row, max_col = dims

    col_type_cache: dict[int, Optional[str]] = {}

    compiled: list[dict[str, Any]] = []
    for i, s in enumerate(parsed_sorts):
        column_id = s["columnId"]
        direction = s["direction"]
        col_idx = resolve_column_id(
            ws,
            selected_sheet=selected_sheet,
            header_row_idx=header_row_idx,
            column_id=column_id,
        )
        inferred = _infer_column_type_cached(
            ws, header_row_idx=header_row_idx, col_idx=col_idx, cache=col_type_cache
        )
        compiled.append(
            {
                "columnId": column_id,
                "direction": direction,
                "colIdx": col_idx,
                "inferredType": inferred,
                "idx": i,
            }
        )

    # Snapshot all data rows before mutating anything.
    row_snaps: list[dict[str, Any]] = []
    for row_idx in range(header_row_idx + 1, max_row + 1):
        cells, dim_snap = _snapshot_row(ws, row_idx=row_idx, max_col=max_col)
        values = [c["value"] for c in cells]
        is_empty = _is_empty_row(tuple(values))
        row_snaps.append(
            {
                "rowIdx": row_idx,
                "cells": cells,
                "dimSnap": dim_snap,
                "values": values,
                "isEmpty": is_empty,
            }
        )

    non_empty_rows = [r for r in row_snaps if not r["isEmpty"]]
    empty_rows = [r for r in row_snaps if r["isEmpty"]]

    if len(non_empty_rows) <= 1:
        return

    # Stable multi-criteria sort: apply from lowest priority to highest.
    for crit in reversed(compiled):
        reverse = crit["direction"] == "desc"
        col_idx = crit["colIdx"]
        inferred_type = crit["inferredType"]
        non_empty_rows.sort(
            key=lambda r, ci=col_idx, it=inferred_type, rev=reverse: _make_sort_key(
                r["values"][ci - 1] if (ci - 1) < len(r["values"]) else None,
                inferred_type=it,
                reverse=rev,
            ),
            reverse=reverse,
        )

    final_rows = non_empty_rows + empty_rows

    for offset, snap in enumerate(final_rows):
        dest_row_idx = header_row_idx + 1 + offset
        _apply_row_snapshot(
            ws,
            dest_row_idx=dest_row_idx,
            cells=snap["cells"],
            dim_snap=snap["dimSnap"],
        )

