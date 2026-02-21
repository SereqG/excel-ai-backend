from __future__ import annotations

from copy import copy

from openpyxl import Workbook, load_workbook


def extract_selected_sheet_workbook(*, original_path: str, selected_sheet: str) -> Workbook:
    """
    Load workbook from original_path and return a NEW workbook containing only selected_sheet.

    Copies values + basic styles similarly to `file_manager.tasks.process_spreadsheet_sheet`.
    """
    wb = load_workbook(original_path, read_only=False, data_only=True)

    if selected_sheet not in wb.sheetnames:
        raise ValueError(
            f'Sheet "{selected_sheet}" not found in workbook. '
            f'Available sheets: {", ".join(wb.sheetnames)}'
        )

    new_wb = Workbook()
    new_wb.remove(new_wb.active)  # remove default sheet

    source = wb[selected_sheet]
    target = new_wb.create_sheet(title=selected_sheet)

    for row in source.iter_rows():
        for cell in row:
            new_cell = target[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                if cell.font:
                    new_cell.font = copy(cell.font)
                if cell.border:
                    new_cell.border = copy(cell.border)
                if cell.fill:
                    new_cell.fill = copy(cell.fill)
                if cell.number_format:
                    new_cell.number_format = cell.number_format
                if cell.protection:
                    new_cell.protection = copy(cell.protection)
                if cell.alignment:
                    new_cell.alignment = copy(cell.alignment)

    return new_wb

