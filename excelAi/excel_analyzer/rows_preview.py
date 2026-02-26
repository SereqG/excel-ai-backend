from __future__ import annotations

import random
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from .services import _extract_column_names, _find_header_row_index, _is_empty_row


def parse_truthy_query_flag(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def coerce_cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def make_unique_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for h in headers:
        base = h if h else "Column"
        count = seen.get(base, 0)
        if count == 0:
            result.append(base)
        else:
            result.append(f"{base}_{count + 1}")
        seen[base] = count + 1
    return result


def row_to_dict(*, row_index: int, row: tuple[Any, ...], headers: list[str]) -> dict[str, Any]:
    payload: dict[str, Any] = {"__rowIndex": row_index}
    for i, header in enumerate(headers):
        payload[header] = coerce_cell_value(row[i] if i < len(row) else None)
    return payload


def _iter_non_empty_data_rows(sheet, *, header_row_idx: int):
    start_row_idx = header_row_idx + 1
    for row_index, row in enumerate(
        sheet.iter_rows(min_row=start_row_idx, values_only=True), start=start_row_idx
    ):
        if _is_empty_row(row):
            continue
        yield row_index, row


def _take_first_rows(
    sheet, *, header_row_idx: int, limit: int
) -> list[tuple[int, tuple[Any, ...]]]:
    rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_index, row in _iter_non_empty_data_rows(sheet, header_row_idx=header_row_idx):
        rows.append((row_index, row))
        if len(rows) >= limit:
            break
    return rows


def _reservoir_sample_rows(
    sheet, *, header_row_idx: int, limit: int
) -> list[tuple[int, tuple[Any, ...]]]:
    reservoir: list[tuple[int, tuple[Any, ...]]] = []
    seen = 0
    for row_index, row in _iter_non_empty_data_rows(sheet, header_row_idx=header_row_idx):
        seen += 1
        if len(reservoir) < limit:
            reservoir.append((row_index, row))
            continue

        j = random.randint(1, seen)
        if j <= limit:
            reservoir[j - 1] = (row_index, row)

    reservoir.sort(key=lambda x: x[0])
    return reservoir


def preview_rows_for_processed_file(
    *,
    original_path: str,
    selected_sheet: str,
    random_sample: bool,
    limit: int = 10,
) -> dict[str, Any]:
    wb = load_workbook(original_path, read_only=True, data_only=True)
    try:
        if selected_sheet not in wb.sheetnames:
            raise ValueError(
                f'Sheet "{selected_sheet}" not found in workbook. '
                f'Available sheets: {", ".join(wb.sheetnames)}'
            )

        sheet = wb[selected_sheet]
        header_row_idx = _find_header_row_index(sheet)
        if header_row_idx is None:
            return {"sheet_name": selected_sheet, "header": [], "rows": [], "total_rows": 0}

        header_row = next(
            sheet.iter_rows(
                min_row=header_row_idx, max_row=header_row_idx, values_only=True
            )
        )
        headers = make_unique_headers(_extract_column_names(header_row))

        picked = (
            _reservoir_sample_rows(sheet, header_row_idx=header_row_idx, limit=limit)
            if random_sample
            else _take_first_rows(sheet, header_row_idx=header_row_idx, limit=limit)
        )
        rows = [row_to_dict(row_index=idx, row=row, headers=headers) for idx, row in picked]

        return {
            "sheet_name": selected_sheet,
            "header": headers,
            "rows": rows,
            "total_rows": len(rows),
        }
    finally:
        wb.close()

