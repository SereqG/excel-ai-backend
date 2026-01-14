"""
Business logic for Excel file analysis.
"""
from openpyxl import load_workbook
from typing import Dict, Any, Union, List, Optional
from django.core.files.uploadedfile import UploadedFile
from datetime import datetime


def _filter_non_empty_values(values: List[Any]) -> List[Any]:
    """
    Filter out None and empty string values from a list.
    
    Args:
        values: List of cell values
        
    Returns:
        List of non-empty values
    """
    return [v for v in values if v is not None and v != '']


def _is_numeric_value(value: Any) -> bool:
    """
    Check if a value is numeric (int, float, or numeric string).
    
    Args:
        value: Value to check
        
    Returns:
        True if value is numeric, False otherwise
    """
    if isinstance(value, (int, float)):
        return True
    
    if isinstance(value, str):
        try:
            float(value.replace(',', ''))
            return True
        except (ValueError, AttributeError):
            return False
    
    return False


def _is_date_value(value: Any) -> bool:
    """
    Check if a value is a date (datetime object or date string).
    
    Args:
        value: Value to check
        
    Returns:
        True if value is a date, False otherwise
    """
    if isinstance(value, datetime):
        return True
    
    if isinstance(value, str):
        date_formats = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']
        for date_format in date_formats:
            try:
                datetime.strptime(value, date_format)
                return True
            except (ValueError, TypeError):
                continue
    
    return False


def _count_numeric_values(values: List[Any]) -> int:
    """
    Count how many values in the list are numeric.
    
    Args:
        values: List of values to check
        
    Returns:
        Number of numeric values
    """
    count = 0
    for value in values:
        if _is_numeric_value(value):
            count += 1
    return count


def _count_date_values(values: List[Any]) -> int:
    """
    Count how many values in the list are dates.
    
    Args:
        values: List of values to check
        
    Returns:
        Number of date values
    """
    count = 0
    for value in values:
        if _is_date_value(value):
            count += 1
    return count


def _count_boolean_values(values: List[Any]) -> int:
    """
    Count how many values in the list are booleans.
    
    Args:
        values: List of values to check
        
    Returns:
        Number of boolean values
    """
    return sum(1 for v in values if isinstance(v, bool))


def _determine_type_from_counts(
    total: int,
    bool_count: int,
    date_count: int,
    number_count: int
) -> str:
    """
    Determine column type based on value type counts.
    
    Args:
        total: Total number of non-empty values
        bool_count: Number of boolean values
        date_count: Number of date values
        number_count: Number of numeric values
        
    Returns:
        Column type ('string', 'number', 'date', 'boolean')
    """
    if bool_count == total:
        return 'boolean'
    
    if date_count / total >= 0.5:
        return 'date'
    
    if number_count / total >= 0.8:
        return 'number'
    
    return 'string'


def _determine_column_type(values: List[Any]) -> Optional[str]:
    """
    Determine the type of a column based on its values.
    
    Args:
        values: List of cell values from the column
        
    Returns:
        Column type ('string', 'number', 'date', 'boolean') or None if cannot determine
    """
    if not values:
        return None
    
    non_empty_values = _filter_non_empty_values(values)
    
    if not non_empty_values:
        return None
    
    bool_count = _count_boolean_values(non_empty_values)
    date_count = _count_date_values(non_empty_values)
    number_count = _count_numeric_values(non_empty_values)
    total = len(non_empty_values)
    
    return _determine_type_from_counts(total, bool_count, date_count, number_count)


def _is_empty_row(row: tuple) -> bool:
    """
    Check if a row is empty (all cells are None or empty).
    
    Args:
        row: Tuple of cell values
        
    Returns:
        True if row is empty, False otherwise
    """
    return all(cell is None or cell == '' for cell in row)


def _get_sample_data(values: List[Any], max_samples: int = 5) -> List[Any]:
    """
    Get sample data from a column, filtering out None and empty values.
    
    Args:
        values: List of cell values from the column
        max_samples: Maximum number of samples to return
        
    Returns:
        List of sample values (converted to JSON-serializable format)
    """
    # Filter out None and empty strings
    non_empty_values = [v for v in values if v is not None and v != '']
    
    # Take up to max_samples
    samples = non_empty_values[:max_samples]
    
    # Convert to JSON-serializable format
    result = []
    for sample in samples:
        if isinstance(sample, datetime):
            # Convert datetime to ISO format string
            result.append(sample.isoformat())
        elif isinstance(sample, (int, float, bool, str)):
            result.append(sample)
        else:
            # Convert other types to string
            result.append(str(sample))
    
    return result


def _find_header_row_index(sheet) -> Optional[int]:
    """
    Find the index of the first non-empty row in a sheet (header row).
    
    Args:
        sheet: OpenPyXL worksheet object
        
    Returns:
        Row index (1-based) of the header row, or None if sheet is empty
    """
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not _is_empty_row(row):
            return idx
    return None


def _extract_column_names(header_row: tuple) -> List[str]:
    """
    Extract column names from the header row.
    
    Args:
        header_row: Tuple of cell values from the header row
        
    Returns:
        List of column names (with default names for empty cells)
    """
    column_names = []
    for i, cell in enumerate(header_row):
        if cell is not None:
            column_name = str(cell)
        else:
            column_name = f'Column_{i+1}'
        column_names.append(column_name)
    return column_names


def _collect_column_data(sheet, header_row_idx: int, num_columns: int) -> Dict[int, List[Any]]:
    """
    Collect data for each column from the sheet, skipping empty rows.
    
    Args:
        sheet: OpenPyXL worksheet object
        header_row_idx: Index of the header row (1-based)
        num_columns: Number of columns to collect data for
        
    Returns:
        Dictionary mapping column index to list of values
    """
    column_data = {i: [] for i in range(num_columns)}
    
    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if _is_empty_row(row):
            continue
        
        for col_idx, cell_value in enumerate(row):
            if col_idx < num_columns:
                column_data[col_idx].append(cell_value)
    
    return column_data


def _process_columns(column_names: List[str], column_data: Dict[int, List[Any]]) -> List[Dict[str, Any]]:
    """
    Process columns to determine types and extract sample data.
    
    Args:
        column_names: List of column names
        column_data: Dictionary mapping column index to list of values
        
    Returns:
        List of column dictionaries with name, type, and sample_data
    """
    columns = []
    for col_idx, col_name in enumerate(column_names):
        column_values = column_data.get(col_idx, [])
        column_type = _determine_column_type(column_values)
        sample_data = _get_sample_data(column_values)
        
        columns.append({
            'name': col_name,
            'type': column_type,
            'sample_data': sample_data
        })
    
    return columns


def _process_sheet(sheet, sheet_name: str) -> Dict[str, Any]:
    """
    Process a single sheet to extract column information.
    
    Args:
        sheet: OpenPyXL worksheet object
        sheet_name: Name of the sheet
        
    Returns:
        Dictionary with sheet name and columns
    """
    header_row_idx = _find_header_row_index(sheet)
    
    if header_row_idx is None:
        return {
            'name': sheet_name,
            'columns': []
        }
    
    header_row = list(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    column_names = _extract_column_names(header_row)
    column_data = _collect_column_data(sheet, header_row_idx, len(column_names))
    columns = _process_columns(column_names, column_data)
    
    return {
        'name': sheet_name,
        'columns': columns
    }


def list_excel_sheets(file: UploadedFile) -> Dict[str, Any]:
    """
    Read an Excel file and return columns with types for each sheet.
    
    Args:
        file: Django uploaded file object
        
    Returns:
        Dictionary containing sheets with their columns and types
    """
    workbook = load_workbook(file, read_only=True, data_only=True)
    
    sheets = []
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_info = _process_sheet(sheet, sheet_name)
        sheets.append(sheet_info)
    
    return {
        'sheets': sheets,
        'total_sheets': len(sheets)
    }
