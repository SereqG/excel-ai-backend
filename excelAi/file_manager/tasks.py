"""
Celery tasks for file_manager app.
"""
import os
import logging
from copy import copy
from celery import shared_task
from django.conf import settings
from django.core.files.base import ContentFile
from openpyxl import load_workbook, Workbook
from .models import ProcessedFile

logger = logging.getLogger(__name__)


@shared_task(bind=True, max_retries=3)
def process_spreadsheet_sheet(self, file_id: str):
    """
    Process a spreadsheet file by extracting the selected sheet.
    
    Args:
        file_id: UUID of the ProcessedFile instance
    """
    try:
        logger.info(f'Processing file {file_id}')
        file_record = ProcessedFile.objects.get(file_id=file_id)
        
        file_record.status = 'PROCESSING'
        file_record.save(update_fields=['status'])
        
        original_path = file_record.original_file.path
        workbook = load_workbook(original_path, read_only=False, data_only=True)
        
        if file_record.selected_sheet not in workbook.sheetnames:
            raise ValueError(
                f'Sheet "{file_record.selected_sheet}" not found in workbook. '
                f'Available sheets: {", ".join(workbook.sheetnames)}'
            )
        
        new_workbook = Workbook()
        new_workbook.remove(new_workbook.active)  # Remove default sheet
        
        # Copy the selected sheet
        source_sheet = workbook[file_record.selected_sheet]
        new_sheet = new_workbook.create_sheet(title=file_record.selected_sheet)
        
        # Copy all cells from source to destination
        for row in source_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                if cell.has_style:
                    # Copy styles properly by creating new style objects
                    if cell.font:
                        new_cell.font = copy(cell.font)
                    if cell.border:
                        new_cell.border = copy(cell.border)
                    if cell.fill:
                        new_cell.fill = copy(cell.fill)
                    if cell.number_format:
                        new_cell.number_format = cell.number_format
                    if cell.protection:
                        new_cell.protection = copy(cell.protection)
                    if cell.alignment:
                        new_cell.alignment = copy(cell.alignment)
        
        # Save the processed workbook to a temporary location
        processed_filename = f'{file_id}_{file_record.selected_sheet}.xlsx'
        processed_path = os.path.join(
            settings.MEDIA_ROOT,
            'temp',
            'processed',
            processed_filename
        )
        
        # Ensure directory exists
        os.makedirs(os.path.dirname(processed_path), exist_ok=True)
        
        # Save workbook
        new_workbook.save(processed_path)
        
        # Update the file record with processed file
        with open(processed_path, 'rb') as f:
            file_record.processed_file.save(
                processed_filename,
                ContentFile(f.read()),
                save=False
            )
        
        # Mark as ready
        file_record.status = 'READY'
        file_record.save(update_fields=['status', 'processed_file'])
        
        logger.info(f'Successfully processed file {file_id}')
        
    except ProcessedFile.DoesNotExist:
        logger.error(f'File {file_id} not found')
        raise
    except Exception as exc:
        logger.error(f'Error processing file {file_id}: {str(exc)}', exc_info=True)
        
        # Mark as failed
        try:
            file_record = ProcessedFile.objects.get(file_id=file_id)
            file_record.status = 'FAILED'
            file_record.save(update_fields=['status'])
        except ProcessedFile.DoesNotExist:
            pass
        
        # Retry if we haven't exceeded max retries
        raise self.retry(exc=exc, countdown=60)


@shared_task
def cleanup_expired_files():
    """
    Periodic task to clean up expired files.
    Runs every hour via Celery Beat.
    """
    from django.utils import timezone
    from .services import delete_file_from_disk
    
    now = timezone.now()
    expired_files = ProcessedFile.objects.filter(expires_at__lt=now)
    
    deleted_count = 0
    for file_record in expired_files:
        # Delete original file
        if file_record.original_file:
            delete_file_from_disk(file_record.original_file.path)
        
        # Delete processed file
        if file_record.processed_file:
            delete_file_from_disk(file_record.processed_file.path)
        
        # Delete database record
        file_record.delete()
        deleted_count += 1
    
    logger.info(f'Cleaned up {deleted_count} expired files')
    return deleted_count
